"""
clean.py — Preprocessing for Lausanne Housing Visualization (COM-480)
Reads all 6 raw .xlsx files and outputs 6 clean long-format CSVs.

Output files (in data/processed/):
  pop_origin.csv       — population by sex + origin (1979–2024/25), city + district level
  pop_age.csv          — population by age class (1979–2024), city + district level
  households.csv       — households by size (2012–2025), city + district level
  employment.csv       — employment by sex + FTE (2011–2023), city + district level
  dwellings_rooms.csv  — dwellings by number of rooms (2010–2024), city + district level
  dwellings_surface.csv— dwellings by surface area (2010–2024), city + district level

Usage:
  python preprocessing/clean.py
  (run from the repo root, expects data/raw/ and writes data/processed/)
"""

import re
import os
import pandas as pd
import openpyxl
from pathlib import Path

RAW = Path("data/raw")
OUT = Path("data/processed")
OUT.mkdir(parents=True, exist_ok=True)

# ── helper: detect if a cell value is a district/sector label ─────────────────
DISTRICT_RE = re.compile(r"^\d{1,4}\s*[-–]")

def is_city_or_district(val):
    """Returns True for 'Ensemble de la ville' or '1 - Centre' style rows."""
    if not isinstance(val, str):
        return False
    return val.strip().startswith("Ensemble") or bool(DISTRICT_RE.match(val.strip()))

def clean_label(val):
    """Normalise district labels."""
    val = val.strip()
    if val.startswith("Ensemble"):
        return "Ville de Lausanne"
    return val

# ══════════════════════════════════════════════════════════════════════════════
# 1. POPULATION BY SEX + ORIGIN  (QH01_01_01 — 1979–2024)
#    One sheet per year. Columns: Total | Swiss | Foreign | Men | Swiss M |
#    Foreign M | Women | Swiss F | Foreign F
# ══════════════════════════════════════════════════════════════════════════════
def parse_pop_origin():
    path = RAW / "QH01_01_01_Population_totale_selon_le_sexe_et_l_origine_2024.xlsx"
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)

    records = []

    for sheet_name in wb.sheetnames:
        if sheet_name in ("Index", "Feuil2"):
            continue

        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # ── detect layout ──────────────────────────────────────────────
        # NEW layout (2015–2024, sheet '2024-2015'):
        #   col 0=district, 1=total, 2=swiss, 3=foreign,
        #   4=men_total, 5=men_swiss, 6=men_foreign,
        #   7=women_total, 8=women_swiss, 9=women_foreign
        # OLD layout (1979–2014, individual year sheets):
        #   col 0=district, 1=total, 2=swiss, 3=foreign,
        #   col 4=None, 5=men_total, 6=men_swiss, 7=men_foreign,
        #   col 8=None, 9=women_total, 10=women_swiss, 11=women_foreign
        is_new_layout = (sheet_name == "2024-2015")

        current_year = None
        in_data      = False

        for i, row in enumerate(rows):
            cell0 = str(row[0]) if row[0] is not None else ""

            # Detect year header
            year_match = re.search(r"(\d{4})", cell0)
            if "origine" in cell0.lower() and year_match:
                current_year = int(year_match.group(1))
                in_data = False
                continue

            # Detect data-start: row where col 1 == 'Total' (sub-header)
            if row[1] in ("Total", "total") and current_year is not None:
                in_data = True
                continue

            if in_data and is_city_or_district(row[0]):
                try:
                    if is_new_layout:
                        records.append({
                            "year":              current_year,
                            "district":          clean_label(str(row[0])),
                            "pop_total":         row[1],
                            "pop_swiss":         row[2],
                            "pop_foreign":       row[3],
                            "pop_men_total":     row[4],
                            "pop_men_swiss":     row[5],
                            "pop_men_foreign":   row[6],
                            "pop_women_total":   row[7],
                            "pop_women_swiss":   row[8],
                            "pop_women_foreign": row[9],
                        })
                    else:
                        # old layout: None gap at cols 4 and 8
                        records.append({
                            "year":              current_year,
                            "district":          clean_label(str(row[0])),
                            "pop_total":         row[1],
                            "pop_swiss":         row[2],
                            "pop_foreign":       row[3],
                            "pop_men_total":     row[5],
                            "pop_men_swiss":     row[6],
                            "pop_men_foreign":   row[7],
                            "pop_women_total":   row[9]  if len(row) > 9  else None,
                            "pop_women_swiss":   row[10] if len(row) > 10 else None,
                            "pop_women_foreign": row[11] if len(row) > 11 else None,
                        })
                except Exception:
                    pass

    wb.close()

    # Also grab 2025 from Q01_01 file
    # Layout: col0=district, col1=total, col2=swiss, col3=foreign,
    #         col4=men_total, col5=men_swiss, col6=men_foreign,
    #         col7=women_total, col8=women_swiss, col9=women_foreign
    path25 = RAW / "Q01_01_Etat_et_structure_de_la_population_2025.xlsx"
    wb25 = openpyxl.load_workbook(path25, read_only=True, data_only=True)
    ws25 = wb25["QT01.01.01"]
    rows25 = list(ws25.iter_rows(values_only=True))
    in_data = False
    for row in rows25:
        if row[1] in ("Total", "total") and row[2] in ("Suisse", "Suisses"):
            in_data = True; continue
        if in_data and is_city_or_district(row[0]):
            records.append({
                "year":              2025,
                "district":          clean_label(str(row[0])),
                "pop_total":         row[1],
                "pop_swiss":         row[2],
                "pop_foreign":       row[3],
                "pop_men_total":     row[4],
                "pop_men_swiss":     row[5],
                "pop_men_foreign":   row[6],
                "pop_women_total":   row[7],
                "pop_women_swiss":   row[8],
                "pop_women_foreign": row[9] if len(row) > 9 else None,
            })
    wb25.close()

    df = pd.DataFrame(records)
    df["pop_total"] = pd.to_numeric(df["pop_total"], errors="coerce")
    df = df[df["pop_total"] > 0].copy()
    df = df.sort_values(["year", "district"]).reset_index(drop=True)

    # Derived columns
    df["pct_foreign"] = (pd.to_numeric(df["pop_foreign"], errors="coerce") / df["pop_total"]).round(4)
    df["pct_swiss"]   = (pd.to_numeric(df["pop_swiss"],   errors="coerce") / df["pop_total"]).round(4)

    df.to_csv(OUT / "pop_origin.csv", index=False)
    print(f"  pop_origin.csv        — {len(df)} rows, years {df.year.min()}–{df.year.max()}")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# 2. POPULATION BY AGE CLASS  (QH01_01_02 — 1979–2024)
#    Paired sheets (e.g. '2024-2015'). Columns: Total | 0-19 | 20-39 | 40-64 | 65-79 | 80+
# ══════════════════════════════════════════════════════════════════════════════
def parse_pop_age():
    path = RAW / "QH01_01_02_Population_totale_par_classe_d_âge_2024.xlsx"
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)

    records = []

    def add_record(year, row, col_offset):
        """Append one record from a data row starting at col_offset."""
        if not is_city_or_district(row[0]):
            return
        try:
            records.append({
                "year":       year,
                "district":   clean_label(str(row[0])),
                "pop_total":  row[col_offset],
                "age_0_19":   row[col_offset + 1],
                "age_20_39":  row[col_offset + 2],
                "age_40_64":  row[col_offset + 3],
                "age_65_79":  row[col_offset + 4],
                "age_80plus": row[col_offset + 5] if len(row) > col_offset + 5 else None,
            })
        except Exception:
            pass

    for sheet_name in wb.sheetnames:
        if sheet_name == "Index":
            continue

        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # ── Sheet '2024-2015': modern stacked layout ──────────────────
        if sheet_name == "2024-2015":
            current_year = None
            in_data      = False
            for row in rows:
                cell0 = str(row[0]) if row[0] else ""
                ym = re.search(r"(\d{4})", cell0)
                if ("groupe" in cell0.lower() or "classe" in cell0.lower()) and ym:
                    current_year = int(ym.group(1)); in_data = False; continue
                if row[1] in ("Total", "total") and current_year:
                    in_data = True; continue
                if in_data and is_city_or_district(row[0]):
                    add_record(current_year, row, 1)   # col 1 = Total
            continue

        # ── Older paired sheets (e.g. '2013-2014', '1979-1980') ───────
        # Year labels appear in one row (col 6 and col 13), data rows follow.
        # Two side-by-side tables: year_left (cols 1-6) | year_right (cols 8-13)
        year_left = year_right = None
        in_data = False
        for row in rows:
            # Detect year label row: numeric values in col 5 or 6 area
            # (e.g. row 5: None,None,...,2013,None,...,2014,...)
            years_found = [v for v in row if isinstance(v, int) and 1979 <= v <= 2025]
            if len(years_found) >= 2:
                year_left, year_right = years_found[0], years_found[1]
                continue
            elif len(years_found) == 1 and year_left is None:
                year_left = years_found[0]
                continue

            # Detect header row: col 1 == 'Total'
            if row[1] in ("Total", "total") and year_left:
                in_data = True; continue

            if in_data and is_city_or_district(row[0]):
                if year_left:  add_record(year_left,  row, 1)   # left table
                if year_right: add_record(year_right, row, 8)   # right table

    wb.close()

    df = pd.DataFrame(records)
    df["pop_total"] = pd.to_numeric(df["pop_total"], errors="coerce")
    df = df[df["pop_total"] > 0].copy()
    df = df.sort_values(["year", "district"]).reset_index(drop=True)

    for col, label in [("age_0_19","pct_0_19"),("age_20_39","pct_20_39"),
                       ("age_40_64","pct_40_64"),("age_65_79","pct_65_79"),
                       ("age_80plus","pct_80plus")]:
        df[label] = (pd.to_numeric(df[col], errors="coerce") / df["pop_total"]).round(4)

    df.to_csv(OUT / "pop_age.csv", index=False)
    print(f"  pop_age.csv           — {len(df)} rows, years {df.year.min()}–{df.year.max()}")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# 3. HOUSEHOLDS BY SIZE  (Q01_03 — 2012–2025)
#    One sheet per year, named by year. Columns: Total|1|2|3|4|5|6+|mean size
# ══════════════════════════════════════════════════════════════════════════════
def parse_households():
    path = RAW / "Q01_03_Ménages_2025.xlsx"
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)

    records = []
    for sheet_name in wb.sheetnames:
        if sheet_name == "Index":
            continue
        try:
            year = int(sheet_name)
        except ValueError:
            continue

        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        in_data = False
        for row in rows:
            # Header: row where col 1 is 'Total' (households count)
            if row[1] in ("Total", "total"):
                in_data = True; continue

            if in_data and is_city_or_district(row[0]):
                try:
                    records.append({
                        "year":        year,
                        "district":    clean_label(str(row[0])),
                        "hh_total":    row[1],
                        "hh_1person":  row[2],
                        "hh_2persons": row[3],
                        "hh_3persons": row[4],
                        "hh_4persons": row[5],
                        "hh_5persons": row[6],
                        "hh_6plus":    row[7],
                        "hh_mean_size":row[8],
                    })
                except Exception:
                    pass

    wb.close()

    df = pd.DataFrame(records)
    df["hh_total"] = pd.to_numeric(df["hh_total"], errors="coerce")
    df = df[df["hh_total"] > 0].copy()
    df = df.sort_values(["year", "district"]).reset_index(drop=True)

    for c in ["hh_1person","hh_2persons"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    df["pct_1person"] = (df["hh_1person"] / df["hh_total"]).round(4)
    df["pct_small_hh"] = ((df["hh_1person"] + df["hh_2persons"]) / df["hh_total"]).round(4)

    df.to_csv(OUT / "households.csv", index=False)
    print(f"  households.csv        — {len(df)} rows, years {df.year.min()}–{df.year.max()}")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# 4. EMPLOYMENT BY SEX  (Q03_01 — 2011–2023)
#    All years stacked in sheet Q03.01.01. Blocks separated by year header rows.
# ══════════════════════════════════════════════════════════════════════════════
def parse_employment():
    path = RAW / "Q03_01_Vie_active_2023.xlsx"
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb["Q03.01.01"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    records = []
    current_year = None
    in_data      = False

    for i, row in enumerate(rows):
        cell0 = str(row[0]) if row[0] is not None else ""

        # New year block header
        year_match = re.search(r"(\d{4})", cell0)
        if "sexe" in cell0.lower() and year_match:
            current_year = int(year_match.group(1))
            in_data = False
            continue

        # Sub-header row: col 1 == 'Etablissements'
        if row[1] in ("Etablissements", "etablissements"):
            in_data = True
            continue
        # Skip the second sub-header row (Total/Hommes/Femmes)
        if in_data and row[2] in ("Total", "total"):
            continue

        if in_data and is_city_or_district(row[0]):
            try:
                records.append({
                    "year":                current_year,
                    "district":            clean_label(str(row[0])),
                    "establishments":      row[1],
                    "jobs_total":          row[2],
                    "jobs_men":            row[3],
                    "jobs_women":          row[4],
                    "fte_total":           row[5],
                    "fte_men":             row[6],
                    "fte_women":           row[7],
                })
            except Exception:
                pass
        elif in_data and row[0] is None and row[1] is None:
            pass  # allow blank rows within a block

    df = pd.DataFrame(records)
    for c in ["jobs_total","jobs_women","fte_total","fte_women"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df = df[df["jobs_total"] > 0].copy()
    df = df.sort_values(["year", "district"]).reset_index(drop=True)

    df["pct_jobs_women"]  = (df["jobs_women"] / df["jobs_total"]).round(4)
    df["pct_fte_women"]   = (df["fte_women"]  / df["fte_total"]).round(4)
    df["gender_fte_gap"]  = (df["pct_jobs_women"] - df["pct_fte_women"]).round(4)

    df.to_csv(OUT / "employment.csv", index=False)
    print(f"  employment.csv        — {len(df)} rows, years {df.year.min()}–{df.year.max()}")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# 5. DWELLINGS BY NUMBER OF ROOMS  (Q09_01 sheet QT09.01.01 — 2010–2024)
#    All years stacked in same sheet, separated by year-header rows.
# ══════════════════════════════════════════════════════════════════════════════
def parse_dwellings_rooms():
    path = RAW / "Q09_01_Construction_logement_2024_new.xlsx"
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb["QT09.01.01"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    records = []
    current_year = None
    in_data      = False

    for row in rows:
        cell0 = str(row[0]) if row[0] is not None else ""

        year_match = re.search(r"(\d{4})", cell0)
        if "nombre de pièces" in cell0.lower() and year_match:
            current_year = int(year_match.group(1))
            in_data = False
            continue

        if row[1] in ("Total", "total") and current_year is not None:
            in_data = True; continue

        if in_data and is_city_or_district(row[0]):
            try:
                records.append({
                    "year":        current_year,
                    "district":    clean_label(str(row[0])),
                    "dw_total":    row[1],
                    "dw_1room":    row[2],
                    "dw_2rooms":   row[3],
                    "dw_3rooms":   row[4],
                    "dw_4rooms":   row[5],
                    "dw_5rooms":   row[6],
                    "dw_6plus":    row[7],
                })
            except Exception:
                pass
        elif in_data and row[0] is None and row[1] is None:
            pass

    df = pd.DataFrame(records)
    for c in ["dw_total","dw_1room","dw_2rooms","dw_5rooms","dw_6plus"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    df = df[df["dw_total"] > 0].copy()
    df = df.sort_values(["year", "district"]).reset_index(drop=True)

    df["pct_small_dw"]  = ((df["dw_1room"] + df["dw_2rooms"]) / df["dw_total"]).round(4)
    df["pct_large_dw"]  = ((df["dw_5rooms"] + df["dw_6plus"]) / df["dw_total"]).round(4)

    df.to_csv(OUT / "dwellings_rooms.csv", index=False)
    print(f"  dwellings_rooms.csv   — {len(df)} rows, years {df.year.min()}–{df.year.max()}")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# 6. DWELLINGS BY SURFACE AREA  (Q09_01 sheet QT09.01.05 — 2010–2024)
#    Same stacked structure as rooms sheet.
# ══════════════════════════════════════════════════════════════════════════════
def parse_dwellings_surface():
    path = RAW / "Q09_01_Construction_logement_2024_new.xlsx"
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb["QT09.01.05"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    records = []
    current_year = None
    in_data      = False

    for row in rows:
        cell0 = str(row[0]) if row[0] is not None else ""

        year_match = re.search(r"(\d{4})", cell0)
        if "surface" in cell0.lower() and year_match:
            current_year = int(year_match.group(1))
            in_data = False
            continue

        if row[1] in ("Total", "total") and current_year is not None:
            in_data = True; continue

        if in_data and is_city_or_district(row[0]):
            try:
                records.append({
                    "year":         current_year,
                    "district":     clean_label(str(row[0])),
                    "dw_total":     row[1],
                    "dw_lt40":      row[2],   # < 40 m2
                    "dw_40_59":     row[3],   # 40–59 m2
                    "dw_60_79":     row[4],   # 60–79 m2
                    "dw_80_99":     row[5],   # 80–99 m2
                    "dw_100_119":   row[6],   # 100–119 m2
                    "dw_120_159":   row[7],   # 120–159 m2
                    "dw_160plus":   row[8],   # 160 m2+
                })
            except Exception:
                pass

    df = pd.DataFrame(records)
    for c in ["dw_total","dw_lt40","dw_40_59","dw_120_159","dw_160plus"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    df = df[df["dw_total"] > 0].copy()
    df = df.sort_values(["year", "district"]).reset_index(drop=True)

    df["pct_small_surface"] = ((df["dw_lt40"] + df["dw_40_59"]) / df["dw_total"]).round(4)
    df["pct_large_surface"] = ((df["dw_120_159"] + df["dw_160plus"]) / df["dw_total"]).round(4)

    df.to_csv(OUT / "dwellings_surface.csv", index=False)
    print(f"  dwellings_surface.csv — {len(df)} rows, years {df.year.min()}–{df.year.max()}")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# RUN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("Preprocessing Lausanne datasets...\n")
    parse_pop_origin()
    parse_pop_age()
    parse_households()
    parse_employment()
    parse_dwellings_rooms()
    parse_dwellings_surface()
    print("\nAll done. CSVs written to data/processed/")